import openpyxl
from openpyxl import load_workbook, Workbook
import os


def write_to_excel(pd):
    """
    Append a row to an existing Excel file or create a new one if it doesn't exist.
    : pd is proc_data dataclass
    """
    filename = epdata_path / "excel" / f"{today.isoformat}.xlxs"

    e_surname = pd.endoscopist.split()[-1]

    data_row = [
        pd.in_theatre,
        e_surname,
        pd.full_name,
        pd.upper,
        pdcolon,
        pd.polyp,
        pd.banding,
    ]

    try:
        # Check if file exists
        if os.path.exists(filename):
            # Load existing workbook
            workbook = load_workbook(filename)
            sheet = workbook.active  # Uses the active sheet
        else:
            # Create a new workbook if file doesn't exist
            workbook = Workbook()
            sheet = workbook.active
            # Optionally add headers if it's a new file
            # sheet.append(['Column1', 'Column2', 'Column3'])  # Customize headers as needed

        # Append the new row
        if not "test" in pd.message:  #  ? exclude if Mr Test File
            sheet.append(data_row)

        # Save the workbook
        workbook.save(filename)
        print(f"Data appended to {filename} successfully!")

    except PermissionError:
        print(
            f"Error: Unable to write to {filename}. File might be open in another program."
        )
    except Exception as e:
        print(f"An error occurred: {e}")


# Example usage
def main():
    # File path for your Excel file
    excel_file = "medical_data.xlsx"

    # Different ways to add rows
    append_to_excel(excel_file, ["John Doe", 45, "Checkup"])
    append_to_excel(excel_file, ["Jane Smith", 38, "Blood Test"])

    # You can mix data types
    mixed_data_row = [
        "Patient Name",  # string
        datetime.now(),  # datetime object
        3.14,  # float
        True,  # boolean
    ]
    write_to_excel(excel_file, mixed_data_row)


# Optional: Add datetime import if using datetime
from datetime import datetime

# Run the main function
if __name__ == "__main__":
    main()
